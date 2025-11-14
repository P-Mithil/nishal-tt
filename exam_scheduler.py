"""Exam timetable scheduler for generating exam schedules."""
import pandas as pd
import random
from file_manager import FileManager
from excel_loader import ExcelLoader
from config import TARGET_SEMESTERS, DEPARTMENTS, PRE_MID, POST_MID, DAYS

class ExamScheduler:
    """Handles scheduling of exams for Pre-Mid and Post-Mid courses."""
    
    def __init__(self, data_frames, schedule_generator):
        self.dfs = data_frames
        self.schedule_gen = schedule_generator
    
    def get_all_pre_mid_courses(self):
        """Get all Pre-Mid courses from all target semesters."""
        all_pre_mid = []
        
        for semester in TARGET_SEMESTERS:
            # Get all semester courses
            sem_courses_all = ExcelLoader.get_semester_courses(self.dfs, semester)
            if sem_courses_all.empty:
                continue
            
            # Parse LTPSC
            sem_courses_parsed = ExcelLoader.parse_ltpsc(sem_courses_all)
            if sem_courses_parsed.empty:
                continue
            
            # Get Pre-Mid courses for each department
            for department in DEPARTMENTS:
                # Filter for department
                if 'Department' in sem_courses_parsed.columns:
                    dept_mask = sem_courses_parsed['Department'].astype(str).str.contains(f"^{department}$", na=False, regex=True)
                    dept_courses = sem_courses_parsed[dept_mask].copy()
                else:
                    dept_courses = sem_courses_parsed.copy()
                
                if dept_courses.empty:
                    continue
                
                # Divide by session
                pre_mid_courses, post_mid_courses = ExcelLoader.divide_courses_by_session(
                    dept_courses, department, all_sem_courses=sem_courses_parsed
                )
                
                # Add Pre-Mid courses (with semester and department info)
                if not pre_mid_courses.empty:
                    pre_mid_courses = pre_mid_courses.copy()
                    pre_mid_courses['Semester'] = semester
                    pre_mid_courses['Dept_Session'] = f"{department}_{PRE_MID}"
                    all_pre_mid.append(pre_mid_courses)
        
        if not all_pre_mid:
            return pd.DataFrame()
        
        # Combine all Pre-Mid courses
        all_pre_mid_df = pd.concat(all_pre_mid, ignore_index=True)
        
        # Remove duplicates by Course Code (same course in multiple departments)
        if 'Course Code' in all_pre_mid_df.columns:
            all_pre_mid_df = all_pre_mid_df.drop_duplicates(subset=['Course Code'], keep='first').reset_index(drop=True)
        
        return all_pre_mid_df
    
    def get_all_post_mid_courses(self):
        """Get all Post-Mid courses from all target semesters."""
        all_post_mid = []
        
        for semester in TARGET_SEMESTERS:
            # Get all semester courses
            sem_courses_all = ExcelLoader.get_semester_courses(self.dfs, semester)
            if sem_courses_all.empty:
                continue
            
            # Parse LTPSC
            sem_courses_parsed = ExcelLoader.parse_ltpsc(sem_courses_all)
            if sem_courses_parsed.empty:
                continue
            
            # Get Post-Mid courses for each department
            for department in DEPARTMENTS:
                # Filter for department
                if 'Department' in sem_courses_parsed.columns:
                    dept_mask = sem_courses_parsed['Department'].astype(str).str.contains(f"^{department}$", na=False, regex=True)
                    dept_courses = sem_courses_parsed[dept_mask].copy()
                else:
                    dept_courses = sem_courses_parsed.copy()
                
                if dept_courses.empty:
                    continue
                
                # Divide by session
                pre_mid_courses, post_mid_courses = ExcelLoader.divide_courses_by_session(
                    dept_courses, department, all_sem_courses=sem_courses_parsed
                )
                
                # Add Post-Mid courses (with semester and department info)
                if not post_mid_courses.empty:
                    post_mid_courses = post_mid_courses.copy()
                    post_mid_courses['Semester'] = semester
                    post_mid_courses['Dept_Session'] = f"{department}_{POST_MID}"
                    all_post_mid.append(post_mid_courses)
        
        if not all_post_mid:
            return pd.DataFrame()
        
        # Combine all Post-Mid courses
        all_post_mid_df = pd.concat(all_post_mid, ignore_index=True)
        
        # Remove duplicates by Course Code (same course in multiple departments)
        if 'Course Code' in all_post_mid_df.columns:
            all_post_mid_df = all_post_mid_df.drop_duplicates(subset=['Course Code'], keep='first').reset_index(drop=True)
        
        return all_post_mid_df
    
    def schedule_exams(self, courses_df, num_days=7):
        """Schedule exams across specified number of days.
        Creates format with days as columns and all courses distributed.
        Returns two DataFrames: one for FN session, one for AN session."""
        if courses_df.empty:
            return pd.DataFrame(), pd.DataFrame()
        
        # Get unique course codes
        if 'Course Code' not in courses_df.columns:
            return pd.DataFrame(), pd.DataFrame()
        
        course_list = courses_df['Course Code'].dropna().unique().tolist()
        random.shuffle(course_list)  # Randomize for distribution
        
        total_courses = len(course_list)
        if total_courses == 0:
            return pd.DataFrame(), pd.DataFrame()
        
        # Use 7 days: Saturday, Monday, Tuesday, Wednesday, Thursday, Friday, Monday
        exam_days = ['Saturday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Monday']
        exam_days = exam_days[:num_days]
        
        # Initialize schedule for FN and AN sessions
        # Each session is a dictionary: {day: [course1, course2, ...]}
        fn_schedule = {day: [] for day in exam_days}
        an_schedule = {day: [] for day in exam_days}
        
        # Distribute all courses across days and sessions
        # Calculate courses per day (divide total courses across all days)
        courses_per_day_base = total_courses // len(exam_days)
        extra_courses = total_courses % len(exam_days)
        
        course_idx = 0
        
        for day_idx, day in enumerate(exam_days):
            # Calculate courses for this day
            courses_this_day = courses_per_day_base
            # Distribute extra courses to first few days
            if day_idx < extra_courses:
                courses_this_day += 1
            
            # Ensure at least 1 course per day if we have courses remaining
            if course_idx < total_courses and courses_this_day == 0:
                courses_this_day = 1
            
            # Distribute courses between FN and AN sessions
            # Split roughly evenly: slightly more in FN if odd number
            fn_count = (courses_this_day + 1) // 2  # Round up
            an_count = courses_this_day - fn_count
            
            # Assign courses to FN session
            for i in range(fn_count):
                if course_idx < total_courses:
                    fn_schedule[day].append(course_list[course_idx])
                    course_idx += 1
            
            # Assign courses to AN session
            for i in range(an_count):
                if course_idx < total_courses:
                    an_schedule[day].append(course_list[course_idx])
                    course_idx += 1
        
        # Create DataFrames with days as columns
        # FN Session DataFrame
        fn_data = {}
        for day in exam_days:
            courses = fn_schedule.get(day, [])
            if courses:
                fn_data[day] = [', '.join([str(c) for c in courses])]
            else:
                fn_data[day] = ['']
        
        fn_df = pd.DataFrame(fn_data)
        
        # AN Session DataFrame
        an_data = {}
        for day in exam_days:
            courses = an_schedule.get(day, [])
            if courses:
                an_data[day] = [', '.join([str(c) for c in courses])]
            else:
                an_data[day] = ['']
        
        an_df = pd.DataFrame(an_data)
        
        return fn_df, an_df
    
    def _format_worksheet(self, worksheet, has_index=False, start_row=1):
        """Format worksheet to ensure all text is clearly visible."""
        try:
            from openpyxl.styles import Alignment, Font
            from openpyxl.utils import get_column_letter
            
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            if max_row == 0 or max_col == 0:
                return
            
            # Format header row
            header_font = Font(bold=True, size=11)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Calculate column widths
            column_widths = {}
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                
                for row_idx in range(1, max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        cell_value = str(cell.value)
                        max_length = max(max_length, len(cell_value))
                
                # Set appropriate width
                if max_length > 30:
                    column_widths[col_letter] = min(max(20, max_length * 0.6), 50)
                else:
                    column_widths[col_letter] = min(max(15, max_length * 1.2), 35)
            
            # Apply formatting
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    col_letter = get_column_letter(col_idx)
                    
                    if row_idx == start_row:
                        cell.font = header_font
                        cell.alignment = header_alignment
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    if col_letter in column_widths:
                        worksheet.column_dimensions[col_letter].width = column_widths[col_letter]
            
            # Set row heights
            if start_row <= max_row:
                worksheet.row_dimensions[start_row].height = 30
            
            for row_idx in range(start_row + 1, max_row + 1):
                worksheet.row_dimensions[row_idx].height = 25
                
        except Exception as e:
            print(f"    WARNING: Could not format worksheet: {e}")
    
    def _format_exam_worksheet(self, worksheet):
        """Format exam worksheet with FN and AN sections."""
        try:
            from openpyxl.styles import Alignment, Font
            from openpyxl.utils import get_column_letter
            
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            if max_row == 0 or max_col == 0:
                return
            
            # Format headers (FN and AN section headers)
            header_font = Font(bold=True, size=12)
            section_font = Font(bold=True, size=11)
            cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Calculate column widths
            column_widths = {}
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                
                for row_idx in range(1, max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        cell_value = str(cell.value)
                        max_length = max(max_length, len(cell_value))
                
                # First column (labels) - narrower
                if col_idx == 1:
                    column_widths[col_letter] = max(20, min(max_length * 1.1, 30))
                else:
                    # Day columns - wider for course codes
                    column_widths[col_letter] = max(15, min(max_length * 0.8, 40))
            
            # Apply formatting
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    col_letter = get_column_letter(col_idx)
                    
                    cell_value = str(cell.value) if cell.value is not None else ''
                    
                    # Format section headers (FN and AN rows)
                    if 'FN:' in cell_value or 'AN:' in cell_value:
                        cell.font = section_font
                        cell.alignment = left_alignment
                    # Format day/date headers
                    elif row_idx in [2, 3, 6, 7] and col_idx > 1:
                        cell.font = Font(bold=True, size=10)
                        cell.alignment = cell_alignment
                    # Format course code cells
                    elif row_idx in [4, 8] and col_idx > 1:
                        cell.alignment = cell_alignment
                        # Set appropriate height for wrapped text
                        if cell_value:
                            lines = len(cell_value.split(','))
                            worksheet.row_dimensions[row_idx].height = max(25, 20 * lines)
                    # Format first column (labels)
                    elif col_idx == 1:
                        if 'Course Code' in cell_value:
                            cell.font = Font(bold=True, size=10)
                        cell.alignment = left_alignment
                    else:
                        cell.alignment = cell_alignment
                    
                    # Apply column width
                    if col_letter in column_widths:
                        worksheet.column_dimensions[col_letter].width = column_widths[col_letter]
            
            # Set row heights
            worksheet.row_dimensions[1].height = 25  # FN header
            worksheet.row_dimensions[2].height = 25  # FN Day row
            worksheet.row_dimensions[3].height = 25  # FN Date row
            worksheet.row_dimensions[4].height = 30  # FN Course codes (adjustable)
            worksheet.row_dimensions[5].height = 10  # Separator
            worksheet.row_dimensions[6].height = 25  # AN header
            worksheet.row_dimensions[7].height = 25  # AN Day row
            worksheet.row_dimensions[8].height = 25  # AN Date row
            worksheet.row_dimensions[9].height = 30  # AN Course codes (adjustable)
                
        except Exception as e:
            print(f"    WARNING: Could not format exam worksheet: {e}")
    
    def _create_exam_sheet(self, fn_df, an_df, exam_days):
        """Create a formatted exam sheet with FN and AN sections."""
        from datetime import datetime, timedelta
        import calendar
        
        # Find next Saturday from today (or use a fixed start date)
        # For now, use a sample date range (Sept 20, 2025 onwards)
        start_date = datetime(2025, 9, 20)  # Saturday, Sept 20, 2025
        
        # Create dates for each day
        dates = []
        current_date = start_date
        day_names_short = ['Sat', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Mon']
        day_names_full = ['Saturday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Monday']
        
        day_to_date = {}
        current_date = start_date
        
        for i, day in enumerate(exam_days):
            if i == 0:
                # First day (Saturday)
                day_to_date[day] = start_date
                current_date = start_date
            else:
                # Calculate days to add based on previous day
                prev_day = exam_days[i-1]
                if prev_day == 'Saturday':
                    days_to_add = 2  # Skip Sunday
                elif prev_day == 'Friday':
                    days_to_add = 3  # Skip weekend
                elif prev_day == 'Monday' and i > 1:
                    # If previous Monday was not first, calculate normally
                    days_to_add = 1
                else:
                    days_to_add = 1
                
                current_date += timedelta(days=days_to_add)
                day_to_date[day] = current_date
        
        # Create combined sheet structure
        # Section 1: FN (10:00 AM - 11:30 AM)
        fn_rows = []
        
        # FN Header
        fn_header = ['FN: 10:00 AM to 11:30 AM']
        for day in exam_days:
            fn_header.append('')
        fn_rows.append(fn_header)
        
        # FN Day row
        fn_day_row = ['']
        for day in exam_days:
            fn_day_row.append(day)
        fn_rows.append(fn_day_row)
        
        # FN Date row
        fn_date_row = ['']
        for day in exam_days:
            date = day_to_date[day]
            fn_date_row.append(date.strftime('%d-%m-%Y'))
        fn_rows.append(fn_date_row)
        
        # FN Course codes row
        fn_course_row = ['Course Code']
        for day in exam_days:
            if day in fn_df.columns and not fn_df.empty:
                courses = fn_df[day].iloc[0] if len(fn_df[day]) > 0 else ''
                fn_course_row.append(courses if courses else '')
            else:
                fn_course_row.append('')
        fn_rows.append(fn_course_row)
        
        # Empty row separator
        fn_rows.append([''] * (len(exam_days) + 1))
        
        # Section 2: AN (03:00 PM - 04:30 PM)
        an_rows = []
        
        # AN Header
        an_header = ['AN: 03:00 PM to 04:30 PM']
        for day in exam_days:
            an_header.append('')
        an_rows.append(an_header)
        
        # AN Day row
        an_day_row = ['']
        for day in exam_days:
            an_day_row.append(day)
        an_rows.append(an_day_row)
        
        # AN Date row
        an_date_row = ['']
        for day in exam_days:
            date = day_to_date[day]
            an_date_row.append(date.strftime('%d-%m-%Y'))
        an_rows.append(an_date_row)
        
        # AN Course codes row
        an_course_row = ['Course Code']
        for day in exam_days:
            if day in an_df.columns and not an_df.empty:
                courses = an_df[day].iloc[0] if len(an_df[day]) > 0 else ''
                an_course_row.append(courses if courses else '')
            else:
                an_course_row.append('')
        an_rows.append(an_course_row)
        
        # Combine all rows
        all_rows = fn_rows + an_rows
        
        # Create DataFrame
        # Find max columns
        max_cols = max(len(row) for row in all_rows) if all_rows else len(exam_days) + 1
        
        # Pad rows to same length
        padded_rows = []
        for row in all_rows:
            padded_row = row + [''] * (max_cols - len(row))
            padded_rows.append(padded_row)
        
        # Create column names
        column_names = [''] + [f'Day_{i}' for i in range(1, max_cols)]
        
        df = pd.DataFrame(padded_rows, columns=column_names[:max_cols])
        return df
    
    def export_exam_timetable(self):
        """Export exam timetable to Excel file with FN and AN sections."""
        print("\n" + "="*80)
        print("GENERATING EXAM TIMETABLE")
        print("="*80)
        
        filename = "exam_timetable.xlsx"
        filepath = FileManager.get_output_path(filename)
        
        try:
            # Get all Pre-Mid and Post-Mid courses
            print("\nCollecting Pre-Mid courses...")
            pre_mid_courses = self.get_all_pre_mid_courses()
            print(f"Found {len(pre_mid_courses)} unique Pre-Mid courses")
            
            print("\nCollecting Post-Mid courses...")
            post_mid_courses = self.get_all_post_mid_courses()
            print(f"Found {len(post_mid_courses)} unique Post-Mid courses")
            
            # Schedule exams (returns FN and AN DataFrames)
            print("\nScheduling Mid-Semester exams (Pre-Mid courses)...")
            mid_fn_df, mid_an_df = self.schedule_exams(pre_mid_courses, num_days=7)
            exam_days = ['Saturday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Monday']
            
            print("\nScheduling End-Semester exams (Post-Mid courses)...")
            end_fn_df, end_an_df = self.schedule_exams(post_mid_courses, num_days=7)
            
            # Create Excel writer
            writer = pd.ExcelWriter(filepath, engine='openpyxl')
            
            try:
                with writer as w:
                    # Create Mid-Semester sheet
                    if not mid_fn_df.empty or not mid_an_df.empty:
                        mid_sem_sheet = self._create_exam_sheet(mid_fn_df, mid_an_df, exam_days)
                        mid_sem_sheet.to_excel(w, sheet_name='mid sem', index=False, header=False)
                        print(f"  Written Mid-Semester schedule with {len(pre_mid_courses)} courses")
                    else:
                        # Create empty sheet
                        empty_df = pd.DataFrame(columns=exam_days)
                        empty_df.to_excel(w, sheet_name='mid sem', index=False)
                        print("  WARNING: No Pre-Mid courses found - created empty Mid-Semester sheet")
                    
                    # Create End-Semester sheet
                    if not end_fn_df.empty or not end_an_df.empty:
                        end_sem_sheet = self._create_exam_sheet(end_fn_df, end_an_df, exam_days)
                        end_sem_sheet.to_excel(w, sheet_name='end sem', index=False, header=False)
                        print(f"  Written End-Semester schedule with {len(post_mid_courses)} courses")
                    else:
                        # Create empty sheet
                        empty_df = pd.DataFrame(columns=exam_days)
                        empty_df.to_excel(w, sheet_name='end sem', index=False)
                        print("  WARNING: No Post-Mid courses found - created empty End-Semester sheet")
                    
                    # Format worksheets
                    try:
                        # Format mid sem sheet
                        ws_mid = w.sheets['mid sem']
                        self._format_exam_worksheet(ws_mid)
                        
                        # Format end sem sheet
                        ws_end = w.sheets['end sem']
                        self._format_exam_worksheet(ws_end)
                        
                        print("  Applied formatting to worksheets")
                    except Exception as e:
                        print(f"  WARNING: Could not format worksheets: {e}")
                
                print(f"\nSUCCESS: Created {filename}")
                print(f"  - Mid-Semester sheet: {len(pre_mid_courses)} courses distributed")
                print(f"  - End-Semester sheet: {len(post_mid_courses)} courses distributed")
                print(f"  - File saved in: {FileManager.OUTPUT_DIR}")
                
                return True
                
            except Exception as e:
                print(f"ERROR: Could not write to Excel file: {e}")
                import traceback
                traceback.print_exc()
                return False
                
        except Exception as e:
            print(f"ERROR: Could not create exam timetable: {e}")
            import traceback
            traceback.print_exc()
            return False

